#!/usr/bin/env python3
"""
Script to create an Excel template for bulk upload with the exact columns the app expects.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

def create_excel_template():
    """Create an Excel template with the exact columns needed for bulk upload."""
    
    # Create sample data with all required columns
    sample_data = {
        'ad_group_name': ['PPCL Services', 'PPCL Services', 'PPCL Consulting'],
        'headline1': ['PPCL Services', 'PPCL Experts', 'PPCL Consulting'],
        'headline2': ['Professional Solutions', 'Trusted Advisors', 'Expert Advice'],
        'headline3': ['Contact Us', 'Learn More', 'Contact Us'],
        'headline4': ['Expert PPCL', 'Professional PPCL', 'Professional Help'],
        'headline5': ['Get Started Today', 'Expert Guidance', 'Get Started'],
        'headline6': ['Learn More', 'Quality Solutions', 'Learn More'],
        'headline7': ['Call Now', 'Call Today', 'Call Now'],
        'headline8': ['Free Consultation', 'Free Quote', 'Free Assessment'],
        'headline9': ['Trusted Advisors', 'Experienced Team', 'Trusted Experts'],
        'headline10': ['Quality Service', 'Professional Service', 'Quality Consulting'],
        'headline11': ['Professional Team', 'Best Results', 'Professional Team'],
        'headline12': ['Experienced Staff', 'Reliable Partner', 'Experienced Consultants'],
        'headline13': ['Reliable Service', 'Industry Leaders', 'Reliable Advice'],
        'headline14': ['Best in Industry', 'Contact Now', 'Best Solutions'],
        'headline15': ['Contact Us Now', 'Get Started', 'Contact Today'],
        'description1': ['Get expert PPCL services today', 'Professional PPCL solutions available', 'Get expert PPCL consulting today'],
        'description2': ['Call today for free consultation', 'Expert guidance and support', 'Professional advice and guidance'],
        'description3': ['Professional solutions available', 'Quality service guaranteed', 'Quality consulting services'],
        'description4': ['Trusted advisors ready to help', 'Best results in the industry', 'Trusted experts ready to help'],
        'final_url': ['https://ppcl.com', 'https://ppcl.com', 'https://ppcl.com'],
        'final_url_path': ['/services', '/consulting', '/contact'],
        'keywords': ['PPCL services;professional services;consultation;expert PPCL', 'PPCL experts;professional advice;guidance;trusted advisors', 'PPCL consulting;expert advice;professional help;consulting services'],
        'headline_positions': ['1;2;3;4;5;6;7;8;9;10;11;12;13;14;15', '1;2;3;4;5;6;7;8;9;10;11;12;13;14;15', '1;2;3;4;5;6;7;8;9;10;11;12;13;14;15'],
        'description_positions': ['1;2;3;4', '1;2;3;4', '1;2;3;4']
    }
    
    # Create DataFrame
    df = pd.DataFrame(sample_data)
    
    # Create Excel file with formatting
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Upload Template"
    
    # Add headers
    headers = list(df.columns)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Add sample data
    for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
    
    # Add instructions sheet
    ws2 = wb.create_sheet("Instructions")
    
    instructions = [
        ["BULK UPLOAD TEMPLATE INSTRUCTIONS"],
        [""],
        ["REQUIRED COLUMNS:"],
        ["• ad_group_name - Name of the ad group"],
        ["• headline1 to headline15 - Ad headlines (up to 30 characters each)"],
        ["• description1 to description4 - Ad descriptions (up to 60 characters each)"],
        ["• final_url - Base landing page URL (e.g., https://example.com)"],
        ["• final_url_path - URL path to append (e.g., /services, /contact, /about)"],
        ["• keywords - Keywords separated by semicolons (e.g., 'keyword1;keyword2')"],
        [""],
        ["OPTIONAL COLUMNS:"],
        ["• headline_positions - Headline positions separated by semicolons"],
        ["• description_positions - Description positions separated by semicolons"],
        [""],
        ["URL CONSTRUCTION:"],
        ["• Final URL = final_url + final_url_path"],
        ["• Example: final_url='https://ppcl.com' + final_url_path='/services' = 'https://ppcl.com/services'"],
        ["• If final_url_path is empty, only final_url is used"],
        [""],
        ["IMPORTANT NOTES:"],
        ["• Budget is set at campaign level (not in this file)"],
        ["• All rows go to one campaign (specified in the app)"],
        ["• Keywords support match types: keyword (broad), 'keyword' (phrase), [keyword] (exact)"],
        ["• Each row creates one ad"],
        ["• Keywords are applied per ad group (uses first row's keywords)"],
        [""],
        ["EXAMPLE KEYWORDS:"],
        ["• Broad match: PPCL services"],
        ["• Phrase match: 'PPCL services'"],
        ["• Exact match: [PPCL services]"],
        [""],
        ["FILE FORMAT:"],
        ["• Save as .xlsx or .csv"],
        ["• UTF-8 encoding recommended"],
        [""],
        ["USAGE:"],
        ["1. Fill in your data in the 'Bulk Upload Template' sheet"],
        ["2. Save the file"],
        ["3. Upload in the app's Bulk Upload tab"],
        ["4. Set campaign name and monthly budget in the app"]
    ]
    
    for row_idx, instruction in enumerate(instructions, 1):
        cell = ws2.cell(row=row_idx, column=1, value=instruction[0])
        if row_idx == 1:
            cell.font = Font(bold=True, size=14)
        elif "REQUIRED COLUMNS:" in instruction[0] or "OPTIONAL COLUMNS:" in instruction[0] or "IMPORTANT NOTES:" in instruction[0] or "EXAMPLE KEYWORDS:" in instruction[0] or "FILE FORMAT:" in instruction[0] or "USAGE:" in instruction[0]:
            cell.font = Font(bold=True)
    
    # Adjust column widths
    for ws in [wb["Bulk Upload Template"], wb["Instructions"]]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the file
    filename = "bulk_upload_template.xlsx"
    wb.save(filename)
    print(f"✅ Excel template created: {filename}")
    print(f"📁 File location: {filename}")
    print(f"🌐 Your app is running at: http://localhost:8534")

if __name__ == "__main__":
    create_excel_template() 